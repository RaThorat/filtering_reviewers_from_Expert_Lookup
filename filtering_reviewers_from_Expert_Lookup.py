# -*- coding: utf-8 -*-
"""
Created on Fri Nov  5 16:16:33 2021

@author: r.thorat
"""

#Preparation of code:
#Importing various modules required for analysis
import pandas._libs.tslibs.base
import numpy as np
import os
os.getcwd()
from tkinter import filedialog
import pandas as pd
from tkinter import*
import re
from nameparser import HumanName

#function to get proposals information
def getproposalinfo():
    global df_pv3
    import_file_path1 = filedialog.askopenfilename()
    df_pv3 = pd.read_excel (import_file_path1)

#function to get suggested reviewers    
def getsuggestedreviewers():
    global df_SR
    import_file_path2 = filedialog.askopenfilename()
    df_SR = pd.read_excel (import_file_path2)
    
#function to get reviewers information from YOUR datawarehouse    
def getdwhreviewers():
    global df_dwh
    import_file_path3 = filedialog.askopenfilename()
    df_dwh = pd.read_excel (import_file_path3)

#function to get Expert Lookup CSV
def getelcsv():
    global df_RL
    import_file_path4 = filedialog.askopenfilename()
    df_RL = pd.read_csv(import_file_path4)

#function to get list of non reviewers
def getnonreviewers():
    global df_nonref
    import_file_path5 = filedialog.askopenfilename()
    df_nonref = pd.read_excel(import_file_path5)

#Developing user interface with Tkinter:
#Tkinter buildup    
root=Tk()
topframe=Frame(root)
topframe.pack(side=TOP)
bottomframe=Frame(root)
bottomframe.pack(side=BOTTOM)
redbtn=Button(topframe, command=getproposalinfo, text="get proposal info", fg="Red")
redbtn.pack(side=LEFT)
bluebtn=Button(topframe, command=getsuggestedreviewers, text="get suggested reviewers", fg="Blue")
bluebtn.pack(side=LEFT)
greenbtn=Button(topframe, command=getdwhreviewers, text="get datawarehouse reviewers", fg="Green")
greenbtn.pack(side=LEFT)
blackbtn=Button(bottomframe, command=getelcsv, text="get Expert Lookup reviewers", fg="Black")
blackbtn.pack(side=LEFT)
yellowbtn=Button(bottomframe, command=getnonreviewers, text="get non reviewers", fg="Violet")
yellowbtn.pack(side=LEFT)

root.mainloop()

##Step processing reviewers list from Expert Lookup

#replacing non existent values with nan
df_RL =df_RL.replace(r'^\s*$', np.nan, regex=True)
#resetting index
df_RL=df_RL.reset_index()
grantnumber=df_RL['GrantNumber'][0]
df_pv3=df_pv3[df_pv3['Aanvraag dossier']==grantnumber]
#resetting index
df_pv3=df_pv3.reset_index()
#For getting full name of the applicant from pivot table
df_RL['Hoofdaanvrager']=df_pv3['Hoofdaanvrager']

#Renaming columns
df_RL.rename(columns = {'GrantNumber':'Dossiernummer', 'Name':'Referent', 'Organization':'Inst','Country':'Land',
                        'Scopus Author Detail Page Link': 'ScopusLink', 'Web Search link':'WebSearchLink', 'Notes':'Opmerking'}, inplace=True)
#extracting details about the reviewers, such as title, first name, surname, gender
Title=[]
First_Name=[]
Middle_Name=[]
Last_Name=[]
for person in df_RL['Referent']:
    name = HumanName(person)
    Title.append(name.title)
    First_Name.append(name.first)
    Middle_Name.append(name.middle)
    Last_Name.append(name.last)
df_RL.loc[:, ('Rangorde')] = ""
df_RL.loc[:, ('Titel')]=Title
df_RL.loc[:, ('Voornaam')] = First_Name
df_RL.loc[:, ('Tuss')]=Middle_Name
df_RL.loc[:, ('Achternaam')] = Last_Name
df_RL.loc[:, ('URL')]=''
df_RL.loc[:, ('m/v')]=''
df_RL =df_RL.replace(r'^\s*$', np.nan, regex=True)

#Renaming columns of suggested reviewers
df_SR.rename(columns = {'Grant No.':'Dossiernummer', 'Last Name':'Achternaam', 'First Name':'Voornaam', 'Affiliation':'Inst','Country':'Land', 'Scopus Author ID': 'ScopusLink'}, inplace=True)


#filtering the suggested reviewers for the given proposal
df_SR=df_SR[df_SR['Dossiernummer']==grantnumber]
df_SR=df_SR.reset_index()


# Writing comment in 'opmerking' in the rows of suggested reviewers inserted in the master reviewers list (RL)

Bron=[]#Empty list
for i in range(len(df_RL)):
    for j in range(len(df_SR)):
        if df_RL['Achternaam'][i]==df_SR['Achternaam'][j]:
            item='Aanvrager, geen COI'# if the reviewer is selected by the applicant
        else: item='Expert Lookup'
        Bron.append(item)

#creation of 'Bron' column
df_RL['Bron'] = pd.DataFrame({'col':Bron})


#rearranging columns
df_RL=df_RL[['Rangorde','Bron', 'Dossiernummer', 'Referent','Titel','Voornaam','Tuss','Achternaam', 'Inst', 'Land', 'Email','m/v', 'URL', 'Opmerking', 'ProposalTitle', 'ProposalLink','Applicants','Hoofdaanvrager', 'ScopusLink', 'WebSearchLink', ]]
df_RL=df_RL.reset_index(drop=True)
#df_RL.head(2)


## Step processing suggested reviewers list

df_SR1=df_SR[['Dossiernummer', 'Achternaam', 'Voornaam',
       'ScopusLink', 'Email', 'Inst', 'Land']]
# adding column with constant value
if len(df_SR1.index)>0:
    df_SR1.loc[:, ('Bron')] = 'Aanvrager, maar misschien COI'
    df_SR1.loc[:, ('Rangorde')] = ''
    df_SR1.loc[:, ('Referent')] = ''
    df_SR1.loc[:, ('Titel')]=' '
    df_SR1.loc[:, ('Tuss')]=''
    df_SR1.loc[:, ('URL')]=' '
    df_SR1.loc[:, ('m/v')]=''
    df_SR1.loc[:, ('Opmerking')] = ''
    df_SR1.loc[:, ('ProposalTitle')]=df_RL['ProposalTitle']
    df_SR1.loc[:, ('Referent')]= ''
    df_SR1.loc[:, ('ProposalLink')]=df_RL['ProposalLink']
    df_SR1.loc[:, ('Applicants')]=df_RL['Applicants']
    df_SR1.loc[:, ('Hoofdaanvrager')]=df_RL['Hoofdaanvrager']
    df_SR1.loc[:, ('WebSearchLink')]=''
    #rearranging columns
    df_SR1=df_SR1[['Rangorde','Bron', 'Dossiernummer', 'Referent','Titel','Voornaam','Tuss','Achternaam', 'Inst', 'Land', 'Email','m/v', 'URL', 'Opmerking', 'ProposalTitle', 'ProposalLink','Applicants', 'Hoofdaanvrager','ScopusLink', 'WebSearchLink']]
        
#adding suggested reviewers to reviewers list    
df_RL_ = pd.concat([df_RL, df_SR1], axis=0)

#dropping duplicate records
df_RL_=df_RL_.drop_duplicates(subset=['Dossiernummer', 'Achternaam'],keep="first")

df_RL_=df_RL_.sort_values('Dossiernummer')

df_RL_1=df_RL_.reset_index(drop=True)


# Check if 'the shape of dataframe' referent list has remained the same. This means that no extra columns have been added in the list the source of the reviewers.
# # Step Compare the reference list with pivot table references from data warehouse

# Cleaning: 
#i.columns names change, NeR: Responded negatively, NoR: Not responded, PoR: Responded positively,
#ii. remove unnecessary columns 'NeR2021', 'NeR0', 'NoR2021', 'NoR0','PoR2021','PoR0', 
#iii. filter non reviewers

df_dwh.columns=['Referent','Achternaam', 'Onbeschikbaar', 'Email', 'Inst', 'Geslacht', 'Opmerking', 'Overleden', 'Voornaam', 'Land', 'Dossiernummer', 'Hoofdaanvrager','NeR2022','NeR2021','NeR2020', 'NeR2019','NeR2018', 'NeR2017',
                'NeR2016', 'NeR2015','NoR2022','NoR2021', 'NoR2020', 'NoR2019', 'NoR2018','NoR2017', 'NoR2016','NoR2015','PoR2022', 'PoR2021','PoR2020', 'PoR2019', 'PoR2018', 'PoR2017', 'PoR2016', 'PoR2015', 'Eindtotaal']



df_dwh1 = df_dwh.iloc[6:]
df_dwh1=df_dwh1.reset_index(drop=True)

##Step Find salutation (Dr. or Prof) of the reviewers (Assumption: Only the first salutation of the reviewer is selected.)
df_dwh1.loc[:, ('Titel')] = df_dwh1['Referent'].str.split('.', 1).str[0].str.strip()
df_dwh1.loc[:, ('ScopusLink')]=''
df_dwh1.loc[:, ('WebSearchLink')]=''
df_dwh1.loc[:, ('ProposalTitle')]=''
df_dwh1.loc[:, ('ProposalLink')]=''

#df_dwh1.head()


df_RL2=df_RL_1.merge(df_dwh1, on='Email', how='left', indicator=True)


#replacing NaN values with zero for further calculation
df_RL2['Eindtotaal'].fillna(0, inplace=True)


#changing the type of column Eindtotaal for comparison
np.int64(df_RL2['Eindtotaal'])


#writing in 'opmerking'column whether a reviwer already registered is in the datawarehouse or new to the the datawarehouse  is
df_RL2.loc[df_RL2['Eindtotaal']!=0, 'Opmerking']='Info is in datawarehouse'



#writing in 'opmerking'column whether a reviwer already registered is in the datawarehouse or new to the the datawarehouse  is
df_RL2.loc[df_RL2['Eindtotaal']==0, "Opmerking"]='New'

#Checking if the reviewer present in the database has worked for the funder in year 2021

#replacing NaN values with zero for further calculation
df_RL2['PoR2021'].fillna(0, inplace=True)
df_RL2['NoR2021'].fillna(0, inplace=True)
df_RL2['NeR2021'].fillna(0, inplace=True)


#changing the type of column 2021 for comparison
np.int64(df_RL2['PoR2021'])
np.int64(df_RL2['NoR2021'])
np.int64(df_RL2['NeR2021'])



df_RL2.loc[df_RL2['PoR2021']!=0, 'Opmerking']='Info is in datawarehouse, has assessed another application or round this year (2021)'

df_RL2.loc[df_RL2['NoR2021']!=0, 'Opmerking']='Info is in datawarehouse, did not respond to another request or round this year (2021)'

df_RL2.loc[df_RL2['NoR2021']!=0,'Opmerking']='Info is in datawarehouse, said no to review this year (2021) on another application or round'

df_RL2=df_RL2.drop(columns={'Referent_y', 'Titel_x',
       'Achternaam_y', 'Onbeschikbaar', 'Inst_y', 'm/v', 'Opmerking_y', 'Opmerking_x',
       'Overleden', 'Voornaam_y', 'Land_y', 'Dossiernummer_y',
       'ProposalTitle_y', 'ProposalLink_y', 'Hoofdaanvrager_y',
       'ScopusLink_y', 'WebSearchLink_y',  'NeR2022','NeR2021','NeR2020', 'NeR2019', 'NeR2018',
       'NeR2017', 'NeR2016', 'NeR2015', 'NoR2022','NoR2021', 'NoR2020','NoR2019',
       'NoR2018', 'NoR2017', 'NoR2016', 'NoR2015', 'PoR2022','PoR2021', 'PoR2020',
        'PoR2019', 'PoR2018', 'PoR2017', 'PoR2016', 'PoR2015', 'Eindtotaal',})


df_RL2=df_RL2.rename(columns={'Dossiernummer_x':'Dossiernummer', 'Referent_x':'Referent', 'Titel_y':'Titel',
       'Voornaam_x':'Voornaam', 'Achternaam_x':'Achternaam', 'Inst_x':'Inst', 'Land_x':'Land', 
'ProposalTitle_x':'ProposalTitle', 'ProposalLink_x':'ProposalLink', 'Hoofdaanvrager_x':'Hoofdaanvrager','ScopusLink_x': 'ScopusLink', 'WebSearchLink_x':'WebSearchLink','Geslacht':'m/v'})


df_RL2=df_RL2[['Dossiernummer', 'Bron', 'Referent', 'Titel','Voornaam','Achternaam', 'Inst', 'Land', 'Email','m/v','Opmerking','ProposalTitle', 'ProposalLink', 'Applicants', 'Hoofdaanvrager',
       'ScopusLink', 'WebSearchLink']]


df_RL2 = df_RL2.assign(Rangorde='')
df_RL2 = df_RL2.assign(Tuss='')
df_RL2 = df_RL2.assign(URL='')

df_RL2 =df_RL2.replace(r'^\s*$', np.nan, regex=True)


#changing the order of the columns  
df_RL2 = df_RL2[['Rangorde','Bron','Dossiernummer', 'Referent', 'Titel', 'Voornaam', 'Tuss', 'Achternaam',
       'Inst', 'Land', 'Email', 'Opmerking', 'URL', 'm/v','ProposalTitle', 'ProposalLink', 'Applicants','Hoofdaanvrager',
       'ScopusLink', 'WebSearchLink']]


# # Step For reapplications from 2020, 2019 and from 2018, searching who are those

# Searching for applicants who have previously applied in 2018, 2019 and 2020

#If the applicant from current round year 2021 have applied previous years, the reviewers of those applications can be taken into account while reviewing the current application. 
#The reviewers can be asked to review the application again because they are familiar with the subject. 
#Or those reviewers can be avoided to prevent prejudice against the applications.

#In the funders database, we search for common string of application number for a given year.
#For example, for 2020 the application number usued contains 2020 and the name of the subsidy. 
#That common string is used below to separate the applications for the year 2018, 2019 and 2020.


df_dwh_her=df_dwh1[(df_dwh1['Dossiernummer'].str.contains('COMMON STRING of APPLICATION NUMBERS 2018')) |(df_dwh1['Dossiernummer'].str.contains('COMMON STRING of APPLICATION NUMBERS 2019'))|(df_dwh1['Dossiernummer'].str.contains('COMMON STRING of APPLICATION NUMBERS 2020'))]





df_dwh_her=df_dwh_her.reset_index()


# In[ ]:

df_dwh_her.columns
df_pv3.columns
df_her=df_pv3.merge(df_dwh_her, on='Hoofdaanvrager', how='left', indicator=True)
df_her.columns

#df_pv2 columns: Dossiernummer,Hoofdaanvrager,Hoofdaanvrager_Achternaam,Voornaam,Geslacht,Promotiedatum,C_Email,Correspondentietaal,HoofdOrganisatie,C_Organisatie,C_Postcode,C_Plaats,Titel,Samenvatting,Hoofddiscipline,Subdiscipline,Woord,SubsidieRonde_Naam,atl_Ingetrokken,atl_Gehonoreerd
# In[ ]:


df_her=df_her.drop(columns='Dossiernummer')

df_her=df_her.rename(columns={'Aanvraag dossier':'Dossiernummer', 'Titel_y':'Titel', 'Geslacht':'m/v'})


#Check the columns from datawarehouse database and application information


#To check if in 2020 or 2019 or 2018 there is a reapplication
df_her1=df_her[['Dossiernummer','Referent', 'Titel','Voornaam','Achternaam', 'Inst', 'Land', 'Email','m/v','Opmerking','Hoofdaanvrager', 'NeR2018','NoR2018','PoR2018', 'NeR2019','NoR2019','PoR2019', 'NeR2020','NoR2020','PoR2020']]


df_her1 = df_her1.assign(Rangorde='')
#df_her1 = df_her1.assign(Info_in_datawarehouse='Ja, herindiening')
df_her1 = df_her1.assign(Tuss='')
df_her1 = df_her1.assign(URL='')
df_her1 = df_her1.assign(ProposalTitle='')
df_her1 = df_her1.assign(ProposalLink='')
df_her1 = df_her1.assign(ScopusLink='')
df_her1 = df_her1.assign(WebSearchLink='')


# In[ ]:


i=0
k=df_her1.columns.get_loc('Hoofdaanvrager')
l=len(df_her1.index)
m=len(df_RL2.index)
n=df_RL2.columns.get_loc('Hoofdaanvrager')
for i in range(0,l):
    j=0
    for j in range(0,m):
        if df_her1.iloc[i,k]==df_RL2.iloc[j,n]:
            df_her1['ProposalTitle'][i]=df_RL2['ProposalTitle'][j]
            df_her1['ProposalLink'][i]=df_RL2['ProposalLink'][j]         
#df_her1.head(2)  


# Checking if the reviewer known from the funder database has reviewed applications from year 2018


#replacing NaN values with zero for further calculation
df_her1['PoR2018'].fillna(0, inplace=True)
df_her1['NoR2018'].fillna(0, inplace=True)
df_her1['NeR2018'].fillna(0, inplace=True)


#changing the type of column 2021 for comparison
np.int64(df_her1['PoR2018'])
np.int64(df_her1['NoR2018'])
np.int64(df_her1['NeR2018'])


df_her1.loc[df_her1['PoR2018']==1, 'Opmerking']='Info is in datawarehouse, reviewed an application from the same applicant last time (2018)'



df_her1.loc[df_her1['NoR2018']==1, 'Opmerking']='Info is in datawarehouse, did not respond last time (2018) to review an application from the same applicant'



df_her1.loc[df_her1['NeR2018']==1, 'Opmerking']='Info is in datawarehouse, said no last time (2018) to review an application from the same applicant'

#Checking if the reviewer known from the funder database has reviewed applications from year 2019
#replacing NaN values with zero for further calculation
df_her1['PoR2019'].fillna(0, inplace=True)
df_her1['NoR2019'].fillna(0, inplace=True)
df_her1['NeR2019'].fillna(0, inplace=True)

#changing the type of column 2021 for comparison
np.int64(df_her1['PoR2019'])
np.int64(df_her1['NoR2019'])
np.int64(df_her1['NeR2019'])


df_her1.loc[df_her1['PoR2019']==1, 'Opmerking']='Info is in datawarehouse, reviewed an application from the same applicant last time (2019)'
df_her1.loc[df_her1['NoR2019']==1, 'Opmerking']='Info is in datawarehouse, did not respond last time (2019) to review an application from the same applicant'
df_her1.loc[df_her1['NeR2019']==1, 'Opmerking']='Info is in datawarehouse, said no last time (2019) to review an application from the same applicant'


# Checking if the reviewer known from the funder database has reviewed applications from year 2020
#replacing NaN values with zero for further calculation
df_her1['PoR2020'].fillna(0, inplace=True)
df_her1['NoR2020'].fillna(0, inplace=True)
df_her1['NeR2020'].fillna(0, inplace=True)


#changing the type of column 2020 for comparison
np.int64(df_her1['PoR2020'])
np.int64(df_her1['NoR2020'])
np.int64(df_her1['NeR2020'])


df_her1.loc[df_her1['PoR2020']==1, 'Opmerking']='Info is in datawarehouse, reviewed an application from the same applicant last time (2020)'


df_her1.loc[df_her1['NoR2020']==1, 'Opmerking']='Info is in datawarehouse, did not respond last time (2020) to review an application from the same applicant'


df_her1.loc[df_her1['NeR2020']==1, 'Opmerking']='Info is in datawarehouse, said no last time (2020) to review an application from the same applicant'

df_her1 =df_her1.replace(r'^\s*$', np.nan, regex=True)

#Filtering reviewers who have reviewed ANY proposals of the applicant in past three years
df_her1=df_her1.drop(columns={'NeR2018',
       'NoR2018', 'PoR2018', 'NeR2019', 'NoR2019', 'PoR2019','NeR2020', 'NoR2020', 'PoR2020'})



df_her1.loc[:, ('Bron')]=''
df_her1.loc[:, ('Applicants')]=''


# In[ ]:


#changing the order of the columns  
df_her1 = df_her1[['Rangorde','Bron','Dossiernummer', 'Referent', 'Titel', 'Voornaam', 'Tuss', 'Achternaam',
       'Inst', 'Land', 'Email', 'Opmerking', 'URL', 'm/v','ProposalTitle', 'ProposalLink','Applicants',  'Hoofdaanvrager',
       'ScopusLink', 'WebSearchLink']]


# using the same technique, filtering the reviewers who have reviewed any proposals of the applicant

# In[ ]:


df_dwh2=df_dwh1[(df_dwh1['PoR2018']==1) |(df_dwh1['PoR2019']==1)|(df_dwh1['PoR2020']==1)]
df_dwh2 = df_dwh2.reset_index(drop=True)


# In[ ]:


df_her2=df_pv3.merge(df_dwh2, on='Hoofdaanvrager', how='left', indicator=True)


# In[ ]:


df_her2=df_her2.drop(columns='Dossiernummer')
df_her2=df_her2.rename(columns={'Aanvraag dossier':'Dossiernummer','Titel_y':'Titel', 'Geslacht':'m/v'})
#To check if in 2019 or 2018 there is a reapplication
df_her3=df_her2[['Dossiernummer','Referent', 'Titel','Voornaam','Achternaam', 'Inst', 'Land', 'Email','m/v','Opmerking','Hoofdaanvrager', 'PoR2018', 'PoR2019', 'PoR2020']]
df_her3 = df_her3.assign(Rangorde='')
#df_her1 = df_her1.assign(Info_in_datawarehouse='Ja, herindiening')
df_her3 = df_her3.assign(Tuss='')
df_her3 = df_her3.assign(URL='')
df_her3 = df_her3.assign(ProposalTitle='')
df_her3 = df_her3.assign(ProposalLink='')
df_her3 = df_her3.assign(ScopusLink='')
df_her3 = df_her3.assign(WebSearchLink='')


df_her3.loc[:, ('Opmerking')]='has assessed an application from the main applicant in the past three years'
df_her3.loc[:, ('Bron')]=''
df_her3.loc[:, ('Applicants')]=''


#changing the order of the columns  
df_her3 = df_her3[['Rangorde','Bron','Dossiernummer', 'Referent', 'Titel', 'Voornaam', 'Tuss', 'Achternaam',
       'Inst', 'Land', 'Email', 'Opmerking','URL', 'm/v','ProposalTitle', 'ProposalLink','Applicants',  'Hoofdaanvrager',
       'ScopusLink', 'WebSearchLink']]

#Getting nor-ref for Talent programma as df_nonref

#Filtering non reviewers
#(If an applicant has submitted a list of non reviewers who he/she doesnot want his/her application to sent to review):

#The excel file with the non reviewers name is usually provided per applcation number. 
#Usually non reviwers name is provided with university name and e-mail adress seperated by comma's.
#Seperating nonreviewers' name from the other text, if the name is given in an excel cell with other things such as university name, email adress etc.
df_nonref['referent naam']=df_nonref['Naam'].str.split(',').str[0]
df_nonref.head()
First_Name=[]
Last_Name=[]
for person in df_nonref['referent naam']:
    name = HumanName(person)
    First_Name.append(name.first)
    Last_Name.append(name.last)
df_nonref['First Name'] = First_Name
df_nonref['Last Name'] = Last_Name
#pd.DataFrame({'First_Name':First_Name,'Last_Name':Last_Name})

#Extraction of email adress
#It can be that these non reviewers are present in the list of reviewers from Expert Lookup and funders database.
#They can still stay in the list, but with a note that clearly says that they must not be contacted to review the given application.
#Extraction of email adresses so that we can use the adressess for filtering from reviewers database:
email=[]
for referenten in df_nonref['Naam']:
    # \w matches any non-whitespace character
    # @ for as in the Email
    # + for Repeats a character one or more times
    FindEmail = re.findall("([\w.-]+@[\w.-]+)", referenten)
    # run for loop on the list variable
    for l in FindEmail:
    #set value in email variable
        emaill=l
    #declare local variables to store email addresses
    email.append(emaill)
df_nonref['Email']=email

#extraction of university name
df_uni = pd.DataFrame()
for referenten in df_nonref['Naam']:
    text=referenten
    matchlist = ['Hospital','University','Universitäts','Universität','Università','Hogeschool','Labs', 'Laboratory', 'Zoo','Institute','Institut','School','Ecole','Academy', 'Universite','College','Universitaet,' '* University'] 
    # define all keywords that you need look up
    p = re.compile('^(.*?),\s+(.*?),(.*?)\.')   # regex pattern to match
    # We use a list comprehension using 'any' function to check if any of the item in the matchlist can be found in either group1 or group2 of the pattern match results
    text_match = [m.group(1) if any(x in m.group(1) for x in matchlist) else m.group(2) for m in re.finditer(p,text)]
    df_uni = df_uni.append(text_match, ignore_index=True)
df_nonref['Inst']=df_uni.iloc[0:]
df_nRef=pd.DataFrame(columns=['Proposal Number','Grant No.','Last Name', 'First Name', 'Scopus Author ID', 'OrcId', 'Email', 'Affiliation', 'Country'])
#Extracting last digits from the grant number, to fill in as proposal number. 
#df_nRef['Proposal Number']=df_nonref['Project dossier'].str.split('.').str[-1]
df_nRef['Bron']='Aanvrager'
df_nRef['Dossiernummer']=df_nonref['Dossiernr']
df_nRef['Rangorde']=''
df_nRef['Referent']=df_nonref['referent naam']
df_nRef['Titel']=''
df_nRef['Achternaam']=df_nonref['Last Name']
df_nRef['Tuss']=''
df_nRef['Voornaam']=df_nonref['First Name']   
df_nRef['Email']=df_nonref['Email']
df_nRef['Inst']=df_nonref['Inst']
df_nRef['Land']=''
df_nRef['m/v']=''
df_nRef['Hoofdaanvrager']=df_nonref['Aanvrager']
df_nRef['ProposalTitle']=''
df_nRef['ScopusLink']=''
df_nRef['WebSearchLink']=''
df_nRef['URL']=''
df_nRef['Applicants']=''
df_nRef['Opmerking']='Nonreferent'
df_nRef['ProposalLink']=''
#changing the order of the columns  
df_nRef = df_nRef[['Rangorde','Bron','Dossiernummer', 'Referent', 'Titel', 'Voornaam', 'Tuss', 'Achternaam',
       'Inst', 'Land', 'Email', 'Opmerking','URL', 'm/v','ProposalTitle', 'ProposalLink','Applicants',  'Hoofdaanvrager',
       'ScopusLink', 'WebSearchLink']]  

# Appending the dataframes for from datawarehouse, from reapplicaions last year, from non referentens, from collaborators
#df_RL2=df_RL2.append([df_her1,df_RL2,df_nRef, df_colab])
#df_RL3=pd.DataFrame().append([df_her1,df_RL2,df_nRef,df_colab])
df_RL3=pd.DataFrame().append([df_RL2,df_her1, df_her3, df_nRef])
df_RL3 =df_RL3.replace(r'^\s*$', np.nan, regex=True)


#Dropping the reviewers mentioned multiple times by the proposals
#df_RL1=df_RL1.drop_duplicates(subset=['Email'])
#check
#df_RL1.shape

df_RL3['Bron'] = df_RL3['Bron'].fillna('Expert Lookup')

df_RL3=df_RL3.reset_index(drop=True)

df_RL3 =df_RL3.replace(r'^\s*$', np.nan, regex=True)

df_RL3 = df_RL3.dropna(how='any',
                    subset=['Achternaam'])

#df_RL3=df_RL3.sort_values('Dossiernummer')

df_RL3["Opmerking"].replace({"-": "New"}, inplace=True)
#https://www.kite.com/python/answers/how-to-replace-column-values-in-a-pandas-dataframe-in-python#:~:text=To%20replace%20values%20in%20the,old%20values%20to%20new%20values.


df_RL3=df_RL3.sort_values('Opmerking',ascending=True).drop_duplicates('Email',keep='last')
#Source:https://kanoki.org/2019/10/25/how-to-remove-duplicate-data-from-python-dataframe/
df_RL3=df_RL3.sort_values('Applicants')

df_RL4 = df_RL3.reset_index(drop=True)


# # Stap 9 Export reference list per proposal

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule
import openpyxl
from typing import NoReturn
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

nummer=df_RL_1['Dossiernummer'][0]
df1=df_RL4[df_RL4['Dossiernummer']==nummer]
df1=df1.reset_index(drop=True)
a_list=df1['Applicants'][0]
#Applicants = "|".join(a_list)
titles=df1['ProposalTitle'].unique()
titel=titles[0]

Expert_Lookup_links=df1['ProposalLink'].unique()
Expert_Lookup_link=Expert_Lookup_links[0]


wb = Workbook()

ws = wb.create_sheet("Instructions", 0)
ws['A1']='file no'
ws['A1'].font = Font(bold=True)
ws['B1']= nummer
ws['A2']='Applicants'
ws['A2'].font = Font(bold=True)  
ws['B2']=a_list
ws['A3']='Title'
ws['A3'].font = Font(bold=True)
ws['B3']=titel
#ws['A4']='Herindiening?'
ws['A4'].font = Font(bold=True)
#ws['B4']=herindiening
ws['A5']='Instructions'
ws['A5'].font = Font(bold=True)
ws['B5']='The referees have been filtered automatically on points below. Very occasionally something slips through, so we ask you to check the suggestions.'
ws['B6']='Sometimes the websearch link doesn not work or it grabs the wrong person. Please note that.'
ws['B7']='Sometimes the sponsor does not belong to the application at all in terms of expertise. Please note that.'
ws['B8']='If the first three or four referees do not match the application, it is better not to use the list.'
ws['B9']=''
ws['B10']='The automatically generated suggestions are based on a fingerprint that is generated in Expert Lookup based on the information in the application: the name of the applicant and any co-applicants, the summary and the keywords.'
ws['B11']=''
ws['B12']='The automatically generated suggestions are filtered in "Search and Select" by:'
ws['B13']='a.appropriate expertise for the proposal (based on the above-mentioned fingerprint)'
ws['B14']='b.no involvement in the proposal -> if the sponsor is mentioned in the proposal (for example as a collaboration partner)'
ws['B15']='c.no joint publications (in the last 3 years)'
ws['B16']='d.sufficient seniority (Referees are at least an associate professor or equivalent. If there is an excellent match in expertise, a more junior referee can review.)'
ws['B17']='e.resubmissions: we do not ask the same referees as for the previous submission'
ws['B18']='f.niet als referent gewerkt heeft in de afgelopen 6 maanden voor een andere ronde'
ws['B19']='g.niet als referent gewerkt heeft in de afgelopen drie jaar om een aanvraag te beoordelen van dezelfde HOOFDaanvrager'
ws['B20']=''
ws['B21']='Voor ontbrekende informatie (aanschrijfvolgorde, aanhef, URL, m/v), kun je gebruik maken van de WebSearchLink of de scopus id link.'
ws['B22']=''
ws['B23']='To search for additional referees you can use the Expert lookup link below. Note: you must have an Expert Lookup account to open the link.'
ws['B24']=Expert_Lookup_link
ws['A24']='EL link'
ws['A24'].font = Font(bold=True)

ws['B26'] = 'Work overview referees on the next worksheet (enter the final selection yourself in the data warehouse)'
ws['B26'].font = Font(bold=True)

def columns_best_fit(ws: openpyxl.worksheet.worksheet.Worksheet) -> NoReturn:
    """
    Make all columns best fit
    """
    column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].bestFit = True
columns_best_fit(ws)
df2=df1.drop(['Dossiernummer','Referent','Hoofdaanvrager','Applicants','ProposalTitle','ProposalLink'], axis=1)    
df2=df2.reset_index(drop=True)
df2=df2.sort_values('Email')
ws1 = wb.create_sheet("Referentenlijst", 1) 
for r in dataframe_to_rows(df2, index=False, header=True):
    ws1.append(r)

for cell in ws1['1'] :#+ ws[1]
    cell.style = 'Pandas'
# Formatting Entire Rows
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(fill=red_fill)
r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r.formula = [('$J1="Info is in data warehouse, has assessed another application or round this year (2021)"')]
ws1.conditional_formatting.add("A1:O100", r)
r1 = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r1.formula = [('$J1="has assessed an application from the main applicant in the past three years"')]
ws1.conditional_formatting.add("A1:O100", r1)
r2 = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r2.formula = [('$J1="Info is in data warehouse, reviewed application from same applicant last time (2018)"')]
ws1.conditional_formatting.add("A1:O100", r2)
r3 = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r3.formula = [('$J1="Info is in data warehouse, reviewed application from same applicant last time (2019)"')]
ws1.conditional_formatting.add("A1:O100", r3)
r4 = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r4.formula = [('$J1="Info is in data warehouse, reviewed application from same applicant last time (2020)"')]
ws1.conditional_formatting.add("A1:O100", r4)
r5 = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r5.formula = [('$J1="Nonreviewer"')]
ws1.conditional_formatting.add("A1:O100", r5)
column_widths = []
for row in ws1:
    for i, cell in enumerate(row):
        if len(column_widths) > i:
            if len(str(cell.value)) > column_widths[i]:
                column_widths[i] = len(str(cell.value))
        else:
            column_widths += [len(str(cell.value))]

for i, column_width in enumerate(column_widths):
    ws1.column_dimensions[get_column_letter(i+1)].width = column_width

output_file_name="Reviewers " + str(nummer)+" Output file"+".xlsx"
#saving the excel workbook
wb.save(output_file_name)
